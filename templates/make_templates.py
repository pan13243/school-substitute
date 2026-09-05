# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

wb = openpyxl.Workbook()

# ===== 样式定义 =====
title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
label_font = Font(name='微软雅黑', size=10, bold=True)
note_font = Font(name='微软雅黑', size=9, color='666666')
data_font = Font(name='微软雅黑', size=10)

title_fill = PatternFill("solid", fgColor="4472C4")
header_fill = PatternFill("solid", fgColor="5B9BD5")
light_fill = PatternFill("solid", fgColor="D6E4F0")
period_fill = PatternFill("solid", fgColor="E2EFDA")
teacher_fill = PatternFill("solid", fgColor="FFF2CC")
day_fill = PatternFill("solid", fgColor="DEEBF7")
empty_fill = PatternFill("solid", fgColor="F5F5F5")

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, font=None, fill=None, alignment=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    cell.border = border


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ===== 模板1：总课表模板 =====
ws1 = wb.active
ws1.title = "总课表模板"

set_col_width(ws1, 1, 10)
set_col_width(ws1, 2, 12)
for i in range(3, 23):
    set_col_width(ws1, i, 11)

# 标题
ws1.merge_cells('A1:V1')
c = ws1['A1']
c.value = "总课表模板（必读填写说明在第2~5行）"
style_cell(c, title_font, title_fill, center)

# 说明区
ws1.merge_cells('A2:V2')
c = ws1['A2']
c.value = "【填写说明】请按此格式整理课表后导入系统"
style_cell(c, label_font, light_fill, center)

ws1.merge_cells('A3:V3')
c = ws1['A3']
c.value = "第1行（表头）：填班级名称，如一（1）、二（3）。最多20个班。班级命名格式：年级（一二三四五六）+（序号），全角括号。"
style_cell(c, note_font, empty_fill, left)

ws1.merge_cells('A4:V4')
c = ws1['A4']
c.value = "第2行起（数据区）：每节次占2行——第1行填科目（语文、数学、英语...），第2行填对应教师姓名。节次名称+时间填在每节第1行第1列。"
style_cell(c, note_font, empty_fill, left)

ws1.merge_cells('A5:V5')
c = ws1['A5']
c.value = "示例：第2行A1填「第一节」B1填「8:20-9:00」，A2留空，各班科目填入相应列。第3行各班填对应教师名。"
style_cell(c, note_font, empty_fill, left)

# 班级名称行（第6行）
row = 6
c = ws1.cell(row, 1)
c.value = "节次/班级"
style_cell(c, header_font, header_fill, center)
c = ws1.cell(row, 2)
c.value = "时间"
style_cell(c, header_font, header_fill, center)
classes = ['一（1）', '一（2）', '一（3）', '二（1）', '二（2）', '二（3）',
           '三（1）', '三（2）', '三（3）', '四（1）', '四（2）', '四（3）', '四（4）',
           '五（1）', '五（2）', '五（3）', '五（4）',
           '六（1）', '六（2）', '六（3）', '六（4）']
for i, cls in enumerate(classes):
    c = ws1.cell(row, 3 + i)
    c.value = cls
    style_cell(c, header_font, day_fill, center)

# 节次定义（相对数据起始行 row=7）
periods = [
    ('第一节',   '8:20-9:00'),
    ('第二节',   '9:10-9:50'),
    ('第三节',   '10:30-11:10'),
    ('第四节',   '11:20-12:00'),
    ('第五节',   '14:00-14:40'),
    ('第六节',   '14:50-15:30'),
]

# 每天5块，从 col 3 起，每块 20 列
current_row = 7
day_names = ['星期一', '星期二', '星期三', '星期四', '星期五']
day_fills = [
    PatternFill("solid", fgColor="C5D9F1"),
    PatternFill("solid", fgColor="C6E0B4"),
    PatternFill("solid", fgColor="FFE699"),
    PatternFill("solid", fgColor="F4B084"),
    PatternFill("solid", fgColor="D9D9D9"),
]

for day_idx, day in enumerate(day_names):
    day_start_col = 3 + day_idx * 20
    df = day_fills[day_idx]

    # 星期标题行
    ws1.merge_cells(start_row=current_row, start_column=day_start_col,
                    end_row=current_row, end_column=day_start_col + 19)
    c = ws1.cell(current_row, day_start_col)
    c.value = "【" + day + "】"
    style_cell(c, header_font, df, center)
    current_row += 1

    for (period_name, period_time) in periods:
        # 节次名行
        c = ws1.cell(current_row, day_start_col)
        c.value = period_name
        style_cell(c, label_font, period_fill, center)
        c = ws1.cell(current_row, day_start_col + 1)
        c.value = period_time
        style_cell(c, data_font, period_fill, center)
        for ci in range(20):
            c = ws1.cell(current_row, day_start_col + ci)
            c.value = ""
            style_cell(c, data_font, empty_fill, center)
        current_row += 1

        # 教师行
        for ci in range(20):
            c = ws1.cell(current_row, day_start_col + ci)
            c.value = ""
            style_cell(c, data_font, teacher_fill, center)
        current_row += 1

ws1.freeze_panes = 'A7'
ws1.sheet_view.showGridLines = True


# ===== 模板2：课后服务模板 =====
ws2 = wb.create_sheet("课后服务模板")

set_col_width(ws2, 1, 10)
set_col_width(ws2, 2, 16)
set_col_width(ws2, 3, 14)
for i in range(4, 24):
    set_col_width(ws2, i, 11)

ws2.merge_cells('A1:W1')
c = ws2['A1']
c.value = "课后服务模板（必读填写说明在第2~4行）"
style_cell(c, title_font, title_fill, center)

ws2.merge_cells('A2:W2')
c = ws2['A2']
c.value = "【填写说明】请按此格式整理课后服务安排后导入系统"
style_cell(c, label_font, light_fill, center)

ws2.merge_cells('A3:W3')
c = ws2['A3']
c.value = "第1行（表头）：第A列填「星期」，第B列填「时间段」，第C列填「项目」，第D列起填各班教师姓名。"
style_cell(c, note_font, empty_fill, left)

ws2.merge_cells('A4:W4')
c = ws2['A4']
c.value = "单周/双周教师不同请用「/」分隔（如 张三/李四，表示单周张三、双周李四）。同一教师通用则只填一人姓名。"
style_cell(c, note_font, empty_fill, left)

# 表头（第5行）
row = 5
for col, hdr in enumerate(['星期', '时间段', '项目'], 1):
    c = ws2.cell(row, col)
    c.value = hdr
    style_cell(c, header_font, header_fill, center)

class_cols_after = ['一（1）', '一（2）', '一（3）', '二（1）', '二（2）', '二（3）',
                    '三（1）', '三（2）', '三（3）', '四（1）', '四（2）', '四（3）', '四（4）',
                    '五（1）', '五（2）', '五（3）', '五（4）',
                    '六（1）', '六（2）', '六（3）', '六（4）']
for i, cls in enumerate(class_cols_after):
    c = ws2.cell(row, 4 + i)
    c.value = cls
    style_cell(c, header_font, header_fill, center)

# 示例数据
after_school_data = [
    ('星期一', '15:40-16:20', '课后服务1'),
    ('星期一', '16:25-17:05', '课后服务2'),
    ('星期一', '17:10-17:50', '课后服务3'),
    ('星期二', '15:40-16:20', '课后服务1'),
    ('星期二', '16:25-17:05', '课后服务2'),
    ('星期二', '17:10-17:50', '课后服务3'),
    ('星期三', '15:40-16:20', '课后服务1'),
    ('星期三', '16:25-17:05', '课后服务2'),
    ('星期三', '17:10-17:50', '课后服务3'),
    ('星期四', '15:40-16:20', '课后服务1'),
    ('星期四', '16:25-17:05', '课后服务2'),
    ('星期四', '17:10-17:50', '课后服务3'),
    ('星期四', '15:40-16:20', '社团活动'),
    ('星期四', '16:25-17:05', '社团活动'),
    ('星期五', '14:40-15:20', '课后服务1'),
    ('星期五', '15:25-16:05', '课后服务2'),
    ('星期五', '16:10-16:50', '课后服务3'),
]

row = 6
for (day, time_range, project) in after_school_data:
    c = ws2.cell(row, 1)
    c.value = day
    style_cell(c, data_font, day_fill, center)
    c = ws2.cell(row, 2)
    c.value = time_range
    style_cell(c, data_font, period_fill, center)
    c = ws2.cell(row, 3)
    c.value = project
    style_cell(c, data_font, light_fill, center)
    for ci in range(20):
        c = ws2.cell(row, 4 + ci)
        c.value = ""
        style_cell(c, data_font, empty_fill, center)
    row += 1

ws2.freeze_panes = 'A6'


# ===== 模板3：校历模板 =====
ws3 = wb.create_sheet("校历模板")

set_col_width(ws3, 1, 10)
set_col_width(ws3, 2, 14)
set_col_width(ws3, 3, 14)
set_col_width(ws3, 4, 12)
set_col_width(ws3, 5, 20)

ws3.merge_cells('A1:E1')
c = ws3['A1']
c.value = "校历模板（必读填写说明在第2~3行）"
style_cell(c, title_font, title_fill, center)

ws3.merge_cells('A2:E2')
c = ws3['A2']
c.value = "【填写说明】每行一个周次。周次从第1周（学期第一周，周一）开始，日期格式 YYYY-MM-DD，单双周必填。"
style_cell(c, label_font, light_fill, center)

ws3.merge_cells('A3:E3')
c = ws3['A3']
c.value = "周六/周日需要调休或补课，请在备注中说明（如「周六上课，补周三」）。"
style_cell(c, note_font, empty_fill, left)

# 表头（第4行）
row = 4
for col, hdr in enumerate(['周次', '开始日期', '结束日期', '单双周', '备注'], 1):
    c = ws3.cell(row, col)
    c.value = hdr
    style_cell(c, header_font, header_fill, center)

# 示例数据（21周）
cal_data = []
start = datetime.date(2025, 9, 1)
for w in range(1, 22):
    begin = start + datetime.timedelta(weeks=w - 1)
    end = begin + datetime.timedelta(days=6)
    parity = '单周' if w % 2 == 1 else '双周'
    cal_data.append((f'第{w}周', begin.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'), parity, ''))

row = 5
for (week, begin, end, parity, note) in cal_data:
    vals = [week, begin, end, parity, note]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row, col)
        c.value = val
        if col == 4:
            pf = PatternFill("solid", fgColor="E2EFDA") if parity == '单周' else PatternFill("solid", fgColor="FFF2CC")
            style_cell(c, data_font, pf, center)
        else:
            style_cell(c, data_font, empty_fill, center)
    row += 1

ws3.freeze_panes = 'A5'

# ===== 保存 =====
wb.save('标准导入模板.xlsx')
print("OK: 标准导入模板.xlsx")
